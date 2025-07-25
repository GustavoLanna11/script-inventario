import platform
import socket
import psutil
import os
import getpass
import subprocess
import requests
from openpyxl import Workbook, load_workbook

DATA_DIR = "data"
os.makedirs(DATA_DIR, exist_ok=True)
FILENAME = os.path.join(DATA_DIR, "inventario_maquinas.xlsx")

def get_wmic_value(command):
    try:
        result = subprocess.check_output(command, shell=True)
        lines = result.decode(errors="ignore").strip().split('\n')
        return lines[1].strip() if len(lines) > 1 else "Não encontrado"
    except Exception:
        return "Erro"

def get_windows_name():
    try:
        result = subprocess.check_output(
            ["powershell", "-Command", "(Get-CimInstance Win32_OperatingSystem).Caption"],
            shell=True
        )
        return result.decode(errors="ignore").strip()
    except Exception:
        return "Erro"

def get_windows_license_status():
    try:
        result = subprocess.check_output(
            ['cscript', '//Nologo', 'C:\\Windows\\System32\\slmgr.vbs', '/xpr'],
            shell=True
        )
        output = result.decode(errors="ignore").lower()
        return "Sim" if "permanente" in output or "permanently" in output else "Não"
    except Exception:
        return "Erro"

def get_memory_type():
    try:
        cmd = [
            "powershell",
            "-Command",
            "Get-CimInstance Win32_PhysicalMemory | Select-Object -First 1 -ExpandProperty SMBIOSMemoryType"
        ]
        result = subprocess.check_output(cmd, shell=True)
        code = int(result.decode(errors="ignore").strip())
        memory_types = {
            20: "DDR",
            21: "DDR2",
            22: "DDR2 FB-DIMM",
            24: "DDR3",
            26: "DDR4",
            30: "DDR5"
        }
        return memory_types.get(code, f"Desconhecido (código {code})")
    except Exception:
        return "Erro"

def get_pc_type():
    try:
        result = subprocess.check_output('wmic computersystem get pcSystemType', shell=True)
        lines = result.decode(errors="ignore").strip().split('\n')
        if len(lines) < 2:
            return "Desconhecido"
        code = lines[1].strip()
        pc_types = {
            '1': 'Desktop',
            '2': 'Notebook'
        }
        return pc_types.get(code, 'Outro')
    except Exception:
        return "Erro"

def get_disk_type():
    try:
        cmd = [
            "powershell",
            "-Command",
            "Get-PhysicalDisk | Select-Object -First 1 -ExpandProperty MediaType"
        ]
        result = subprocess.check_output(cmd, shell=True)
        media_type = result.decode(errors="ignore").strip().lower()
        if media_type in ["ssd", "hdd"]:
            return media_type.upper()
        else:
            return "Desconhecido"
    except Exception:
        return "Desconhecido"

def get_city_from_ip():
    try:
        response = requests.get("https://ipinfo.io/json", timeout=5)
        if response.status_code == 200:
            data = response.json()
            return data.get("city", "")
        return ""
    except Exception:
        return ""

def has_kaspersky():
    try:
        cmd = [
            "powershell",
            "-Command",
            "Get-CimInstance -Namespace root/SecurityCenter2 -ClassName AntivirusProduct | Select-Object -ExpandProperty displayName"
        ]
        result = subprocess.check_output(cmd, shell=True)
        output = result.decode(errors="ignore").lower()
        return "Sim" if "kaspersky" in output else "Não"
    except Exception:
        return "Erro"

def get_machine_info():
    info = {}
    processor_name = get_wmic_value("wmic cpu get name").strip()
    weak_cpus = [
        "intel core i3-2120", "intel core i3-3220", "intel core i3-4130",
        "intel core i5-2430m", "amd a4-6300", "intel core i3-4005u",
        "intel core i3-5015u", "intel celeron j1800", "intel celeron g460",
        "intel core i3-3217u"
    ]
    normalized_processor = processor_name.lower().split("@")[0].split("cpu")[0].strip()

    info["Nome da máquina"] = socket.gethostname()
    info["Proprietário"] = getpass.getuser()
    info["Etiqueta"] = ""
    info["Cidade"] = get_city_from_ip()
    info["Departamento"] = ""
    info["Unidade Residente"] = ""
    info["Marca"] = get_wmic_value("wmic computersystem get manufacturer")
    info["Número de Série"] = get_wmic_value("wmic bios get serialnumber")
    info["Tipo"] = get_pc_type()
    info["Modelo"] = get_wmic_value("wmic computersystem get model")
    info["Licença"] = get_windows_name()
    info["Processador"] = processor_name
    info["Tipo de memória"] = get_memory_type()
    info["Pentes"] = "1"

    ram_gb = round(psutil.virtual_memory().total / (1024 ** 3), 2)
    info["Tamanho"] = ram_gb

    disk = psutil.disk_usage('/')
    info["Armazenamento"] = round(disk.total / (1024 ** 3), 2)

    disk_type = get_disk_type()
    info["Tipo de armazenamento"] = disk_type
    print(f"[DEBUG] Tipo de armazenamento identificado: {disk_type}")

    info["Antivírus"] = has_kaspersky()
    info["Em uso?"] = "Sim"
    info["Está no AD?"] = os.environ.get('USERDOMAIN', "")
    info["Observações"] = ""

    is_weak_cpu = any(cpu in normalized_processor for cpu in weak_cpus)

    if is_weak_cpu:
        info["Troca de máquina"] = "Sim"
        info["Upgrade?"] = "Não"
        info["Troca ou Upgrade"] = "Troca"
        info["Prioridade"] = "Alta"
        info["Licença Windows"] = "Máquina para troca"
    else:
        info["Troca de máquina"] = "Não"
        if ram_gb < 4 or disk_type.lower() == "hdd":
            info["Upgrade?"] = "Sim"
            info["Troca ou Upgrade"] = "Upgrade"
            info["Prioridade"] = "Não será trocada"
            info["Licença Windows"] = get_windows_license_status()
        else:
            info["Upgrade?"] = "Não"
            info["Licença Windows"] = get_windows_license_status()
            if info["Troca de máquina"] == "Não" and info["Upgrade?"] == "Não":
                info["Troca ou Upgrade"] = "Nenhum"
            else:
                info["Troca ou Upgrade"] = ""
            info["Prioridade"] = "Não será trocada"

    return info

def save_to_excel(info, filename=FILENAME):
    ordered_keys = [
        "Nome da máquina", "Proprietário", "Etiqueta", "Cidade", "Departamento",
        "Unidade Residente", "Marca", "Número de Série", "Tipo", "Modelo",
        "Licença", "Processador", "Troca de máquina", "Tipo de memória", "Pentes",
        "Tamanho", "Armazenamento", "Tipo de armazenamento", "Licença Windows",
        "Troca ou Upgrade", "Prioridade", "Antivírus", "Upgrade?", "Em uso?",
        "Está no AD?", "Observações"
    ]

    try:
        wb = load_workbook(filename)
        ws = wb.active
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.append(ordered_keys)

    row = [info.get(key, "") for key in ordered_keys]
    ws.append(row)
    wb.save(filename)
    print(f"✅ Planilha '{filename}' salva com sucesso!")

def send_api(filepath):
    try:
        url = "http://192.168.0.138:5000//upload_excel"
        with open(filepath, "rb") as f:
            files = {'file': (os.path.basename(filepath), f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
            response = requests.post(url, files=files)

        if response.status_code == 200:
            print("✅ Arquivo enviado para a API com sucesso.")
        else:
            print(f"⚠️ Erro ao enviar para a API: {response.status_code} - {response.text}")
    except Exception as e:
        print(f"⚠️ Falha ao conectar com a API: {e}")

if __name__ == "__main__":
    info = get_machine_info()
    save_to_excel(info)
    send_api(FILENAME)
